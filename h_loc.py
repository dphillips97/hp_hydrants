import openpyxl
import requests
import json
from api_keys import *

def geocode(loc):	

	API_KEY = GITHUB_API_KEY
	
	URL = r'https://maps.googleapis.com/maps/api/geocode/json?'
	
	city = ',Hazel_Park'
	state = ',MI'

	unique_url = URL + 'address=%s&key=%s' % ((loc + city + state), API_KEY)
	
	results = requests.get(unique_url)
	data = json.loads(results.text)
		
	if (len(data['results'])) != 0:
		
		print('%s - FOUND' % loc)
		
		lat = data['results'][0]['geometry']['location']['lat']
		lng = data['results'][0]['geometry']['location']['lng']
		
		# list only takes 1 argument
		results_tuple = (lat, lng)
		results = list(results_tuple)
	
		return results
	
	else:
		
		print('%s - NOT FOUND' % loc)
		return
	
def main():
	
	# open sheet object
	wb = openpyxl.load_workbook('hydrant_loc.xlsx')
	sheet = wb.active
	
	# get max row and col for later use
	max_row = sheet.max_row
	max_col = sheet.max_column
	
	# for every record after headers
	for row_iter in range(2, max_row + 1):
	
		# get location
		loc = sheet.cell(row = row_iter, column = 2).value
		
		# call geocode function and return list of lat, long
		results = geocode(loc)
		
		# results = [lat, lng]
		if results:
		
			# unpack list
			lat = results[0]
			lng = results[1]
		
			# write values
			sheet.cell(row = row_iter, column = max_col + 1).value = lat
			sheet.cell(row = row_iter, column = max_col + 2).value = lng
			
		else:
		
			pass
		
	wb.save('hydrants_coded.xlsx')
	
		
main()